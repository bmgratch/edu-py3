#! python3
# multiplicationTable.py - creates a spreadsheet multiplication table in Excel
# py multiplicationTable.py (N) - creates a spreadsheet N by N

import openpyxl, sys

# get N from command line
if len(sys.argv) == 2:
    n = sys.argv[1]
else:
    sys.exit()

# Open a sheet
mTable = openpyxl.Workbook()
sheet = mTable.active
sheet.title = 'Multiplication %s' % n

# TODO create the table
# N wide by N down
titleFont = Font(bold=True)
sheet.cell(row = 1, column = 1).value = 'X'
for y in range(1, n + 1):
    sheet.cell(row = y + 1, column = 1).value = y
for x in range(1, n+1):
    sheet.cell(row = 1, column = x + 1).value = x

# Build the multiplication table
for y in range(1, n+1):
    for x in range(1, n + 1):
        sheet.cell(row = y+1, column = x+1).value = x * y

# saving:
mTable.save('multiplication-table-%s.xlsx' % n)
